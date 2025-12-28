import os, sys
import glob
import pandas as pd
import numpy as np
import argparse
import re
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Color

def create_build_dir():
    os.makedirs('./build/', exist_ok=True)
    return os.path.abspath('./build')

def replace_with_regex(build_dir, filename, pattern, replacement):
    # Step 1: Read the entire file content
    print(filename)
    with open(filename, 'r') as file:
        content = file.read()

    # Step 2: Use re.sub() to replace patterns
    # 'pattern' is a regex pattern (e.g., r'\d+' to match any number)
    # 'replacement' is the new string
    modified_content = re.sub(pattern, replacement, content)

    # Step 3: Write the modified content back to the file
    abs_dir = os.path.dirname(filename)
    file_name = os.path.splitext(os.path.basename(filename))[0]
    file_ext = os.path.splitext(os.path.basename(filename))[1]
    
    tmp_file = os.path.join(build_dir, file_name + '_tmp' + file_ext)
    # if os.path.exists(tmp_file):
        # os.remove(tmp_file)
    
    with open(tmp_file, 'w') as file:
        file.write(modified_content)

    return tmp_file

class Chip:
    ref_ptr = 7
    
    def __init__(self, file_path):
        self.__raw_name = ''
        self.__data = None # datetime
        self.__file_path = file_path
        self.__mx = None # np.array 2 dim
        self.__vec = None # np.array 1 dim
        self.__ref_mx = None
        self.__ref_vec = None
        self.__ref_val = 0.0
        self.__ref_err = 0.0
        # self.__data_start = -1
        
        with open(file_path, 'r') as file:
            content = file.readlines()
        
        self.__parse_content(content)
        
    def __parse_content(self, content):
        second_line = content[1].split(' ')
        self.__raw_name = second_line[0]
        
        tmp_data = second_line[1] + second_line[2].replace('\r', '').replace('\n', '')
        self.__data = datetime.datetime.strptime(tmp_data,
                                                 "%d.%m.%Y%H:%M:%S")
        self.__parse_table(content)
        self.__calc_ref_value()
        self.__calc_ref_table()
    
    def __parse_table(self, content):
        start_ptr = self.__find_data_start(content)
        if start_ptr == -1:
            print(f"ERROR: Can't find table of data in {self.__file_path}")
            return
        
        mx = self.__mx
        row = 0
        for line in content[start_ptr:]:
            numbs = re.findall("[-+]?(?:\d*\.*\d+)", line)
            # print(numbs)
            for col, num in enumerate(numbs):
                mx[row, col] = float(num)
            row += 1

        # print(self.__mx)
        self.__vec = self.__mx.reshape(-1, order='F')
        
    def __find_data_start(self, content):
        for n, line in enumerate(content):
            if 'x.y' in line:
                x, y = re.findall(r'\d+', line)
                self.__mx = np.zeros((int(y), int(x)))
                return n + 1
        return -1
    
    def __calc_ref_value(self):
        sum = 0
        cnt = 0
        vec = self.__vec
        
        for v in vec[1:6]:
            sum += v
            cnt += 1
        
        for v in vec[51:55]:
            sum += v
            cnt += 1
        
        self.__ref_val = sum / cnt
        self.__ref_err = vec[self.ref_ptr] / self.__ref_val
    
    def __calc_ref_table(self):
        (n, m) = self.__mx.shape
        size = n*m
        m = m - 1 # remove reference row
        self.__ref_vec = np.zeros((n * m))
        
        start = n
        end = size - n + 2
        # print(start, end)
        for i, v in enumerate(self.__vec[start:end]):
            self.__ref_vec[i] = v / self.__ref_val
        
        self.__ref_mx = self.__ref_vec.reshape((n, -1))
        
        # print(self.__ref_vec) 
        # print('\n')
        print(self.__ref_mx)
    
    def to_excel(self, ws, start_pos):
        name = self.__raw_name
        data = self.__data
        col_ptr = 2
        r = start_pos
        c = 1
        
        # name
        ws.cell(row=r, column=c, value=name)
        
        # data
        ws.cell(row=r+1, column=c, value=data.strftime('%d.%m.%Y'))
        col_ptr += self.__mx_to_excel(ws, start_pos, col_ptr) + 1
        
        self.__ref_mx_to_excel(ws, start_pos, col_ptr)
        
        start_pos = start_pos + self.__mx.shape[0] + 1 # + 1 header of table
        
        return start_pos
    
    def __mx_to_excel(self, ws, row_start, col_start):
        mx = self.__mx
        # (r_max, c_max) = mx.shape
        
        row_start += 1 #
        for idx, val in np.ndenumerate(mx):
            # print(f"{idx}, val: {val}")
            r = idx[0] + row_start
            c = idx[1] + col_start
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.number_format = '0.00'
        
        return c
    
    def __ref_mx_to_excel(self, ws, row_start, col_start):
        mx = self.__ref_mx
        
        for idx, val in np.ndenumerate(mx):
            r = idx[0] + row_start
            c = idx[1] + col_start
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.number_format = '0'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            self.__set_cell_color(cell)
    
    def __set_cell_color(self, cell):
        
        r = '63'
        g = 'BE'
        b = '78'
        color = r + g + b
        print(color)
        cell.fill = PatternFill(bgColor=color, fill_type="solid")
            
parser = argparse.ArgumentParser(description="Parse chip table")
parser.add_argument('-d', '--dir')
namespace = parser.parse_args(sys.argv[1:])

# Suppress Scientific Notation (exponential)
np.set_printoptions(suppress=True)
        
if namespace.dir:
    parse_dir = namespace.dir
else:
    parse_dir = './'

scan_dir = os.path.abspath(parse_dir)

build_dir = create_build_dir()

txtfiles = []
for d,r,f in os.walk(scan_dir):
    for file in f:
        if file.endswith(".txt"):
            txtfiles.append(os.path.join(d, file))

wb = Workbook()
ws = wb.active
ws.title = "snip"

pos = 1
for file in txtfiles:
    modified_file = replace_with_regex(build_dir, file, ',', '.')
    
    chip = Chip(modified_file)
    pos += chip.to_excel(ws, pos) + 3
    # # open file
    # with open(file, 'r') as fd:
        # for line in fd:
wb.save(os.path.join(build_dir, 'b.xlsx'))
wb.close()

# for ws_col in ws.columns:
    # max_len = max((len(str(cell)) for cell in ws_col)) + 1
    # ws.column_dimensions[ws_col[0].column_letter].width = max_len*1.2